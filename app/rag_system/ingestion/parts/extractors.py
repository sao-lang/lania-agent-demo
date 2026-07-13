"""RAG 系统内容提取器子模块。

负责 DOCX/PPTX/CSV/HTML/代码/媒体 等各种格式的内容提取。
与主应用的 `app/rag/ingestion_parts/extractors.py` 功能一致。
"""

from __future__ import annotations

import importlib
import re
import subprocess
from pathlib import Path
from html.parser import HTMLParser
from typing import Any

from app.rag_system.ingestion.parts._typing import RagIngestionTypingMixin


class _RagHTMLTextExtractor(HTMLParser):
    """把 HTML 内容提取为纯文本片段。"""
    def __init__(self):
        super().__init__()
        self.parts: list[str] = []
    def handle_data(self, data: str):
        text = data.strip()
        if text:
            self.parts.append(text)
    def get_text(self) -> str:
        return '\n'.join(self.parts)


class RagIngestionExtractorMixin(RagIngestionTypingMixin):
    """DOCX/PPTX/CSV/LibreOffice/HTML/代码/媒体/ZIP 内容提取。"""

    MIME_OVERRIDES: dict[str, tuple[str, ...] | set[str]] = {}

    def _extract_docx(self, file_path: Path) -> str:
        try:
            docx = importlib.import_module('docx')
            doc = docx.Document(str(file_path))
            return '\n'.join(p.text for p in doc.paragraphs)
        except Exception as exc:
            return f'[DOCX 提取失败: {exc}]'

    def _extract_pptx(self, file_path: Path) -> str:
        try:
            pptx = importlib.import_module('pptx')
            prs = pptx.Presentation(str(file_path))
            texts: list[str] = []
            for slide in prs.slides:
                for shape in slide.shapes:
                    if hasattr(shape, 'text') and shape.text.strip():
                        texts.append(shape.text.strip())
            return '\n'.join(texts)
        except Exception as exc:
            return f'[PPTX 提取失败: {exc}]'

    def _extract_xlsx(self, file_path: Path) -> str:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), data_only=True)
            texts: list[str] = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_text: list[str] = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        rows_text.append(' | '.join(cells))
                if rows_text:
                    texts.append(f'=== {sheet_name} ===\n' + '\n'.join(rows_text[:self.TABLE_SEGMENT_ROW_BATCH]))
            return '\n'.join(texts)
        except Exception as exc:
            return f'[XLSX 提取失败: {exc}]'

    def _extract_csv(self, file_path: Path) -> str:
        try:
            import csv
            texts: list[str] = []
            with open(file_path, newline='', encoding='utf-8', errors='replace') as f:
                reader = csv.reader(f)
                for idx, row in enumerate(reader):
                    if idx > self.TABLE_SEGMENT_ROW_BATCH:
                        break
                    texts.append(' | '.join(row))
            return '\n'.join(texts)
        except Exception as exc:
            return f'[CSV 提取失败: {exc}]'

    def _extract_html(self, file_path: Path) -> str:
        content = file_path.read_text(encoding='utf-8', errors='replace')
        extractor = _RagHTMLTextExtractor()
        extractor.feed(content)
        return extractor.get_text()

    def _convert_with_libreoffice(self, file_path: Path, target_format: str = 'txt') -> str | None:
        try:
            cmd = [self.settings.office_converter_command, '--headless', '--convert-to', target_format, '--outdir', str(self.settings.resolved_data_dir), str(file_path)]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=self.settings.office_conversion_timeout_seconds)
            if result.returncode != 0:
                return None
            converted = Path(self.settings.resolved_data_dir) / f'{file_path.stem}.{target_format}'
            if converted.exists():
                text = converted.read_text(encoding='utf-8', errors='replace')
                converted.unlink(missing_ok=True)
                return text
            return None
        except Exception:
            return None

    def _extract_office_file(self, file_path: Path) -> str:
        ext = file_path.suffix.lstrip('.').lower()
        if ext == 'docx':
            return self._extract_docx(file_path)
        elif ext == 'pptx':
            return self._extract_pptx(file_path)
        elif ext == 'xlsx':
            return self._extract_xlsx(file_path)
        elif ext in ('csv', 'tsv'):
            return self._extract_csv(file_path)
        elif ext in ('htm', 'html'):
            return self._extract_html(file_path)
        else:
            converted = self._convert_with_libreoffice(file_path)
            if converted:
                return converted
            try:
                return file_path.read_text(encoding='utf-8', errors='replace')
            except Exception:
                return f'[无法提取文件: {file_path.name}]'

    def _extract_code_text(self, content: str) -> str:
        return content

    def _extract_archive_member(self, member_path: str, content: bytes) -> str | None:
        ext = member_path.rsplit('.', 1)[-1].lower() if '.' in member_path else ''
        if ext in self.TEXT_LIKE_TYPES:
            return content.decode('utf-8', errors='replace')
        return None
